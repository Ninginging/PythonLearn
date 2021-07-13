import joseph_ring_v2 as v2
import csv      # csv文件模块
from openpyxl import Workbook, load_workbook


source_path = 'Source\\'


class Read2list:
    def __init__(self, file_name, path=source_path):
        self.file_name = file_name
        self.path = path

    def get_list(self):
        file = open(self.path + self.file_name, 'r', encoding='utf-8')
        all_lines = file.readlines()
        file.close()
        results = []
        for i in range(len(all_lines)):
            each_line = all_lines[i].strip('\n')        # 去除换行符
            results.append(each_line.split(','))
        return results


class Read_xls(Read2list):
    def get_list(self):
        xls_file = load_workbook(self.path + self.file_name)
        assert self.file_name.split('.')[1] in ['xls', 'xlsx'], 'file type wrong'
        sheet = xls_file.active
        cols_num = sheet.max_column
        rows_num = sheet.max_row
        results = []
        each_line = []

        for i in range(rows_num):
            for j in range(cols_num):
                each_line.append(sheet.cell(i+1, j+1).value)
            results.append(each_line)
            each_line = []

        return results


# 生成一个txt文件,写入玩家信息
def get_txt(file_name='test_data'):
    txt_file = open(source_path + file_name + '.txt', 'w+', encoding='utf-8')
    for i in range(v2.GROUP_SIZE):
        txt_file.write(v2.group[i].name + ',')
        txt_file.write(v2.group[i].std_id + ',')
        txt_file.write(v2.group[i].gender + ',')
        txt_file.write(v2.group[i].birthdate + '\n')

    txt_file.close()


# 生成一个csv文件,按照txt当中的信息写入
def get_csv(data, file_name='test_data'):
    csv_file = open(source_path + file_name + '.csv', 'w+', encoding='utf-8', newline='')
    writer_csv = csv.writer(csv_file)
    for i in range(len(data)):
        writer_csv.writerow(data[i])
    csv_file.close()


# 生成一个xlsx文件,按照txt当中的信息写入
def get_xls(data, file_name='test_data'):
    work_book = Workbook()
    work_sheet = work_book.active
    for i in range(len(data)):
        for j in range(len(data[0])):
            work_sheet.cell(i+1, j+1).value = data[i][j]
    work_book.save(source_path + file_name + '.xlsx')


if __name__ == '__main__':      # Test
    # get_txt()
    txt_data = Read2list('test_data.txt').get_list()
    # get_csv(txt_data)
    csv_data = Read2list('test_data.csv').get_list()
    # get_xls(txt_data)
    xls_data = Read_xls('test_data.xlsx').get_list()

    group_name, group_name_new = [], []
    for i in range(v2.GROUP_SIZE):
        group_name.append(txt_data[i][0])
        group_name_new.append(v2.Joseph(xls_data, 4).traversal()[i][0])
    print('初始容器顺序为：%s\n遍历之后的结果是：%s' % (group_name, group_name_new))

